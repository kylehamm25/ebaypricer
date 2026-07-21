import argparse
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from ebaypricer.auth import get_access_token
from ebaypricer.trading_api import fetch_sold_orders
from ebaypricer.finances import fetch_finance_fees, merge_fees_into_rows
from ebaypricer.cards import enrich_rows
from ebaypricer.excel import (
    HEADER_FILL, HEADER_FONT, DATA_FONT,
    CURRENCY_COLS, INT_COLS, write_headers,
)

CUTOFF = datetime(2026, 6, 30, tzinfo=timezone.utc)

ORDER_LEVEL_COLS = ("Shipping", "Order Total", "Total eBay Fees", "Order Earnings")

DEFAULT_OUTPUT = r"H:\My Drive\ebay\ebay_sold_orders.xlsx"


def parse_args():
    parser = argparse.ArgumentParser(description="Append new eBay sold orders to existing Excel workbook")
    parser.add_argument("--output", type=str, default=DEFAULT_OUTPUT, help="Output xlsx path")
    parser.add_argument("--days", type=int, default=0,
                        help="Fetch last N days (default 0 = use hardcoded cutoff 2026-06-30)")
    return parser.parse_args()


def read_header_cols(ws: Worksheet) -> dict[str, int]:
    col: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            name = str(cell.value).strip()
            if name not in col:
                col[name] = cell.column - 1
    return col


def get_existing_keys(ws: Worksheet) -> set[tuple[tuple[str, ...], str]]:
    col = read_header_cols(ws)
    date_idx = col.get("Sale Date", 2)
    iid_idx = col.get("Item ID", 1)
    last_data_row = find_last_data_row(ws)
    keys: set[tuple[tuple[str, ...], str]] = set()
    for row in ws.iter_rows(min_row=2, max_row=last_data_row, values_only=True):
        date = str(row[date_idx]).strip() if len(row) > date_idx and row[date_idx] is not None else ""
        iid_raw = str(row[iid_idx]).strip() if len(row) > iid_idx and row[iid_idx] is not None else ""
        item_ids = tuple(sorted(i.strip() for i in iid_raw.split("; ") if i.strip()))
        if item_ids:
            keys.add((item_ids, date))
    return keys


def order_key(order: dict) -> tuple[tuple[str, ...], str]:
    iid_raw = order.get("Item ID", "")
    item_ids = tuple(sorted(i.strip() for i in iid_raw.split("; ") if i.strip()))
    date = order.get("Sale Date", "")
    return (item_ids, date)


def find_last_data_row(ws: Worksheet) -> int:
    for row in range(ws.max_row, 0, -1):
        cell = ws.cell(row=row, column=1)
        if cell.value is not None and str(cell.value).strip():
            return row
    return 1


def create_new_workbook(headers: list[str]) -> tuple[Workbook, Worksheet]:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Sold Orders", 0)
    ws.title = "Sold Orders"
    write_headers(ws, headers)
    ws.freeze_panes = "A2"
    return wb, ws


DEPRECATED_COLS = {
    "Items in Order", "Item Fees (est.)", "Item Earnings (est.)",
    "Card Name", "Card Number", "Set Name", "Set Series",
    "Rarity", "Variant", "Market Price",
}


def _strip_deprecated_cols(ws: Worksheet) -> None:
    col = read_header_cols(ws)
    to_del = sorted(
        (idx for name, idx in col.items() if name in DEPRECATED_COLS),
        reverse=True,
    )
    for idx in to_del:
        ws.delete_cols(idx + 1)


def _deduplicate_headers(ws: Worksheet) -> None:
    seen: set[str] = set()
    to_del: list[int] = []
    for cell in ws[1]:
        if cell.value is not None:
            name = str(cell.value).strip()
            if name in seen:
                to_del.append(cell.column)
            else:
                seen.add(name)
    for col in reversed(to_del):
        ws.delete_cols(col)


def blank_order_level_continuation_rows(rows: list[dict]) -> None:
    groups = defaultdict(list)
    for i, r in enumerate(rows):
        groups[r["Order ID"]].append(i)

    for indices in groups.values():
        if len(indices) <= 1:
            continue
        indices_sorted = sorted(indices, key=lambda i: rows[i].get("Item Price") or 0, reverse=True)
        for i in indices_sorted[1:]:
            for col in ORDER_LEVEL_COLS:
                rows[i][col] = None


def main():
    args = parse_args()
    now = datetime.now(timezone.utc)

    if args.days > 0:
        start_dt = now - timedelta(days=args.days)
        label = start_dt.strftime("%Y-%m-%d")
    else:
        start_dt = CUTOFF
        label = f"{CUTOFF.date()} (hardcoded)"

    token = get_access_token()

    raw_rows = fetch_sold_orders(token, start_dt, now)

    min_date = start_dt.strftime("%Y-%m-%d")
    raw_rows = [r for r in raw_rows if r.get("Sale Date", "") >= min_date]

    seen: dict = {}
    deduped = []
    for r in raw_rows:
        key = (r["Item ID"], r.get("Sale Date", ""))
        if key not in seen:
            seen[key] = len(deduped)
            deduped.append(r)
        else:
            existing = deduped[seen[key]]
            existing_oid = existing.get("Order ID", "")
            candidate_oid = r.get("Order ID", "")
            if existing_oid.startswith(existing.get("Item ID", "")) and not candidate_oid.startswith(r.get("Item ID", "")):
                deduped[seen[key]] = r
    raw_rows = deduped

    if not raw_rows:
        print("No orders found")
        sys.exit(0)

    fee_start = start_dt - timedelta(days=15)
    fees_by_order, item_id_index = fetch_finance_fees(token, fee_start, now)
    merge_fees_into_rows(raw_rows, fees_by_order, item_id_index)

    enrich_rows(raw_rows)

    raw_rows.sort(key=lambda r: (r["Sale Date"], r.get("Buyer") or ""))
    headers = list(raw_rows[0].keys())

    xlsx_path = args.output
    existing_keys: set = set()
    fetched_keys = {order_key(r) for r in raw_rows}

    new_cols: list[str] = []
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb["Sold Orders"]
        _strip_deprecated_cols(ws)
        _deduplicate_headers(ws)
        existing_cols = read_header_cols(ws)
        new_cols = [h for h in headers if h not in existing_cols]
        if new_cols:
            next_col = max(existing_cols.values()) + 2 if existing_cols else 1
            for h in new_cols:
                cell = ws.cell(row=1, column=next_col, value=h)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
                next_col += 1
        existing_keys = get_existing_keys(ws)
    else:
        wb, ws = create_new_workbook(headers)

    new_orders = [r for r in raw_rows if order_key(r) not in existing_keys]
    skipped = len(raw_rows) - len(new_orders)

    if not new_orders:
        if new_cols:
            wb.save(xlsx_path)
        print(f"No new orders")
        sys.exit(0)

    blank_order_level_continuation_rows(new_orders)

    if os.path.exists(xlsx_path):
        start_row = find_last_data_row(ws) + 1
    else:
        start_row = 2

    for row_idx, row in enumerate(new_orders, start_row):
        for h, val in row.items():
            col_idx = read_header_cols(ws).get(h)
            if col_idx is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=val)
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="center")
            if h in CURRENCY_COLS and val is not None:
                cell.number_format = '#,##0.00'
            elif h in INT_COLS:
                cell.number_format = '0'

    wb.save(xlsx_path)
    print(f"Appended {len(new_orders)} ({skipped} dupes)\n")


if __name__ == "__main__":
    main()
