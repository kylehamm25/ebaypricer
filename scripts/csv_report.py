"""
Usage:
    python csv_report.py
    python csv_report.py --days 90
    python csv_report.py --start 2025-01-01 --end 2025-06-30
    python csv_report.py --output my_sales.xlsx
    python csv_report.py --debug-fees
    python csv_report.py --skip-fees
"""

import argparse
import csv
import os
import sys
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ebaypricer.auth import get_access_token, _parse_usd, _parse_csv_date
from ebaypricer.trading_api import fetch_sold_orders
from ebaypricer.finances import fetch_finance_fees, merge_fees_into_rows
from ebaypricer.cards import enrich_rows

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

DEFAULT_OUTPUT = r"H:\My Drive\ebay\ebay_sold_orders.xlsx"


def read_csv_orders(csv_path: str) -> list[dict]:
    if not os.path.isabs(csv_path):
        csv_path = os.path.join(SCRIPT_DIR, csv_path)
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader)
        headers = next(reader)

        col = {h: i for i, h in enumerate(headers)}
        needed = {"Order Number", "Item Number", "Quantity", "Sold For",
                  "Shipping And Handling", "Total Price", "Sale Date",
                  "Buyer Username", "Item Title", "Custom Label"}
        missing = needed - set(col)
        if missing:
            print(f"ERROR: CSV missing columns: {', '.join(sorted(missing))}")
            sys.exit(1)

        rows = []
        summary = {}
        max_col = max(col.values())
        for csv_row in reader:
            if not csv_row or len(csv_row) <= max_col or not csv_row[col["Order Number"]].strip():
                continue
            order_num = csv_row[col["Order Number"]].strip()
            item_num = csv_row[col["Item Number"]].strip()

            if not item_num:
                summary[order_num] = {
                    "shipping": _parse_usd(csv_row[col["Shipping And Handling"]]),
                    "total": _parse_usd(csv_row[col["Total Price"]]),
                }
                continue

            qty = int(csv_row[col["Quantity"]].strip() or "1")
            sold_for = _parse_usd(csv_row[col["Sold For"]])
            shipping = _parse_usd(csv_row[col["Shipping And Handling"]])
            total_price = _parse_usd(csv_row[col["Total Price"]])

            if order_num in summary:
                if shipping == 0.0 and summary[order_num]["shipping"] != 0.0:
                    shipping = summary[order_num]["shipping"]
                if total_price == 0.0 and summary[order_num]["total"] != 0.0:
                    total_price = summary[order_num]["total"]

            rows.append({
                "Order ID": order_num,
                "Item ID": item_num,
                "Sale Date": _parse_csv_date(csv_row[col["Sale Date"]]),
                "Buyer": csv_row[col["Buyer Username"]].strip(),
                "Item Title": csv_row[col["Item Title"]].strip(),
                "Quantity": qty,
                "Item Price": sold_for,
                "Shipping": shipping,
                "Order Total": total_price,
                "SKU": csv_row[col["Custom Label"]].strip(),
            })

    if not rows:
        print("ERROR: No order rows found in CSV.")
        sys.exit(1)

    print(f"Read {len(rows)} line items from CSV")
    return rows


def combine_orders(rows: list[dict]) -> list[dict]:
    groups: dict = {}
    order = []
    for row in rows:
        oid = row["Order ID"]
        if oid not in groups:
            groups[oid] = []
            order.append(oid)
        groups[oid].append(row)

    combined = []
    for oid in order:
        group = groups[oid]
        if len(group) == 1:
            combined.append(group[0])
            continue

        first = group[0]
        skus = [r["SKU"] for r in group if r["SKU"]]
        combined.append({
            "Order ID": oid,
            "Item ID": "; ".join(r["Item ID"] for r in group),
            "Sale Date": first["Sale Date"],
            "Buyer": first["Buyer"],
            "Item Title": "; ".join(r["Item Title"] for r in group),
            "Quantity": sum(r["Quantity"] for r in group),
            "Shipping": first["Shipping"],
            "Order Total": first["Order Total"],
            "Total eBay Fees": first.get("Total eBay Fees"),
            "Order Earnings": first.get("Order Earnings"),
            "SKU": "; ".join(skus),
        })

    print(f"Combined into {len(combined)} orders ({len(rows) - len(combined)} multi-item collapsed)")
    return combined


def write_excel(rows: list[dict], filename: str) -> str:
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Sold Orders"

    headers = list(rows[0].keys()) if rows else []

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    for col_idx, h in enumerate(headers, 1):
        cell: Cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_font = Font(name="Arial", size=10)
    currency_cols = {"Item Price", "Shipping", "Order Total", "Total eBay Fees", "Order Earnings"}
    int_cols = {"Quantity"}

    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            cell: Cell = ws.cell(row=row_idx, column=col_idx, value=row[h])
            cell.font = data_font
            cell.alignment = Alignment(vertical="center")
            if h in currency_cols and row[h] is not None:
                cell.number_format = '#,##0.00'
            elif h in int_cols:
                cell.number_format = '0'

        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                shade: Cell = ws.cell(row=row_idx, column=col_idx)
                shade.fill = PatternFill("solid", fgColor="EBF3FB")

    ws_sum = wb.create_sheet("Summary")
    n = len(rows)
    ws_sum["A1"] = "Summary"
    ws_sum["A1"].font = Font(bold=True, name="Arial", size=12)

    col_letter = {h: get_column_letter(i) for i, h in enumerate(headers, 1)}

    ws_sum["A3"] = "Total Transactions"
    ws_sum["B3"] = f"=COUNTA('Sold Orders'!{col_letter['Order ID']}2:{col_letter['Order ID']}{n + 1})"
    ws_sum["A4"] = "Total Revenue"
    ws_sum["B4"] = f"=SUM('Sold Orders'!{col_letter['Order Total']}2:{col_letter['Order Total']}{n + 1})"
    ws_sum["B4"].number_format = '#,##0.00'
    ws_sum["A5"] = "Total Shipping Collected"
    ws_sum["B5"] = f"=SUM('Sold Orders'!{col_letter['Shipping']}2:{col_letter['Shipping']}{n + 1})"
    ws_sum["B5"].number_format = '#,##0.00'

    next_row = 6
    if "Total eBay Fees" in col_letter:
        ws_sum[f"A{next_row}"] = "Total eBay Fees"
        ws_sum[f"B{next_row}"] = (
            f"=SUM('Sold Orders'!{col_letter['Total eBay Fees']}2:"
            f"{col_letter['Total eBay Fees']}{n + 1})"
        )
        ws_sum[f"B{next_row}"].number_format = '#,##0.00'
        next_row += 1
    if "Order Earnings" in col_letter:
        ws_sum[f"A{next_row}"] = "Total Order Earnings"
        ws_sum[f"B{next_row}"] = (
            f"=SUM('Sold Orders'!{col_letter['Order Earnings']}2:"
            f"{col_letter['Order Earnings']}{n + 1})"
        )
        ws_sum[f"B{next_row}"].number_format = '#,##0.00'
        next_row += 1

    ws_sum[f"A{next_row}"] = "Avg Order Value"
    ws_sum[f"B{next_row}"] = "=IF(B3=0,0,B4/B3)"
    ws_sum[f"B{next_row}"].number_format = '#,##0.00'
    last_row = next_row

    for r in range(3, last_row + 1):
        ws_sum.cell(r, 1).font = Font(name="Arial", size=10, bold=True)
        ws_sum.cell(r, 2).font = Font(name="Arial", size=10)

    note_row = last_row + 2
    notes = []
    if "Order Earnings" in col_letter:
        notes.append(
            "Note: 'Order Earnings' = Order Total minus eBay fees. It does not "
            "subtract refunds (not tracked here) or your own item cost — eBay's "
            "Seller Hub 'Net order earnings' figure additionally subtracts a "
            "per-item cost you enter manually in Seller Hub, which isn't "
            "exposed via any API."
        )
    if any(row.get("Total eBay Fees") is None for row in rows):
        notes.append(
            "Note: some rows show blank fees because no matching transaction "
            "was found in the Finances API response for that order/item within "
            "the date range searched — verify with --debug-fees if unexpected."
        )
    for note in notes:
        ws_sum[f"A{note_row}"] = note
        ws_sum[f"A{note_row}"].font = Font(name="Arial", size=9, italic=True, color="808080")
        note_row += 1

    for col_idx, h in enumerate(headers, 1):
        col_letter_w = get_column_letter(col_idx)
        max_len = max(
            len(str(h)),
            max((len(str(row.get(h, ""))) for row in rows), default=0),
        )
        ws.column_dimensions[col_letter_w].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    wb.save(filename)
    return filename


def parse_args():
    parser = argparse.ArgumentParser(description="Export eBay sold orders to Excel via Trading API")
    parser.add_argument("--csv",    type=str,                                help="Read orders from CSV instead of Trading API")
    parser.add_argument("--days",   type=int, default=30,                    help="Past days to fetch (default: 30)")
    parser.add_argument("--start",  type=str,                                help="Start date YYYY-MM-DD (overrides --days)")
    parser.add_argument("--end",    type=str,                                help="End date YYYY-MM-DD (default: today)")
    parser.add_argument("--output", type=str, default=DEFAULT_OUTPUT, help="Output filename")
    parser.add_argument("--skip-fees", action="store_true", help="Skip the Finances API fee/earnings lookup entirely")
    parser.add_argument("--debug-fees", action="store_true", help="Print raw Finances API JSON samples for verifying field names")
    return parser.parse_args()


def main():
    args = parse_args()

    now    = datetime.now(timezone.utc)
    end_dt = datetime.strptime(args.end, "%Y-%m-%d").replace(tzinfo=timezone.utc) if args.end else now
    start_dt = (
        datetime.strptime(args.start, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        if args.start
        else end_dt - timedelta(days=args.days)
    )

    token = get_access_token()

    if args.csv:
        rows = read_csv_orders(args.csv)
        if not args.start and not args.end:
            dates = [datetime.strptime(r["Sale Date"], "%Y-%m-%d").replace(tzinfo=timezone.utc) for r in rows]
            start_dt = min(dates)
            end_dt = max(dates) + timedelta(days=1)
        print(f"Orders from {start_dt.date()} to {end_dt.date()}")
    else:
        print(f"Fetching sold orders from {start_dt.date()} to {end_dt.date()}...")
        rows = fetch_sold_orders(token, start_dt, end_dt)

    print(f"\nFound {len(rows)} line items.")

    if not rows:
        print("No sold orders found for this date range.")
        sys.exit(0)

    if args.skip_fees:
        for row in rows:
            row["Total eBay Fees"] = None
            row["Order Earnings"] = None
    else:
        print("\nFetching fee/earnings data from Finances API...")
        fees_by_order, item_id_index = fetch_finance_fees(
            token, start_dt, end_dt, debug=args.debug_fees
        )
        print(f"Found fee data for {len(fees_by_order)} orders")
        merge_fees_into_rows(rows, fees_by_order, item_id_index)

    rows = combine_orders(rows)

    print("  Enriching with Pokémon card data ...")
    enrich_rows(rows)
    output_file = write_excel(rows, args.output)
    print(f"Saved -> {output_file}")


if __name__ == "__main__":
    main()
