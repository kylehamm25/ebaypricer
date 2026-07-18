import argparse
import logging
import os
import sqlite3
import sys
import time
from datetime import datetime, timezone
from statistics import mean

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from ebaypricer.browse_api import (
    get_today_active_snapshot,
    init_db,
    parse_active_item,
    save_active_snapshot,
    search_active_listings,
)
from ebaypricer.excel import HEADER_FILL, HEADER_FONT, DATA_FONT
from ebaypricer.paths import DB_PATH as DEFAULT_DB_PATH

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

SHEET_NAME = "Active Listings"
MAX_LISTINGS = 9999
DEFAULT_OUTPUT = r"H:\My Drive\ebay\ebay_sold_orders.xlsx"

ACTIVE_PRICE_COLUMNS = [
    ("Active Avg (Top 5)", '#,##0.00'),
]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Add average active-listing prices to Active Listings sheet"
    )
    parser.add_argument("--output", type=str, default=DEFAULT_OUTPUT,
                        help="Path to the Excel workbook")
    parser.add_argument("--db", type=str, default=DEFAULT_DB_PATH,
                        help="Path to SQLite database")
    parser.add_argument("--max-listings", type=int, default=MAX_LISTINGS,
                        help="Max number of listings to process")
    parser.add_argument("--force", action="store_true",
                        help="Run even if snapshots already exist for today")
    return parser.parse_args()


def read_active_listings(ws) -> tuple[list[dict], list[str], dict[str, int]]:
    headers = [cell.value for cell in ws[1]]
    col_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_map[str(h).strip()] = i

    card_col = col_map.get("Card")
    if card_col is None:
        sys.exit(0)

    rows = []
    for excel_row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) == 0 or row[0] is None:
            continue
        row_dict = {headers[i] if i < len(headers) else None: row[i] for i in range(len(row))}
        row_dict["_row"] = excel_row_num
        rows.append(row_dict)

    return rows, headers, col_map


def collect_unique_cards(rows: list[dict], headers: list[str], col_map: dict[str, int]) -> list[str]:
    seen: set[str] = set()
    cards: list[str] = []
    card_col = col_map.get("Card")
    if card_col is None:
        return cards
    for row in rows:
        val = row.get(headers[card_col]) if card_col < len(headers) else row.get("Card")
        if val and str(val).strip():
            card = str(val).strip()
            if card not in seen:
                seen.add(card)
                cards.append(card)
    return cards


def fetch_active_price_for_card(conn, card_name: str) -> dict | None:
    snapshot = get_today_active_snapshot(conn, card_name)
    if snapshot:
        return snapshot

    print(".", end="", flush=True)
    try:
        items = search_active_listings(card_name)
    except requests.RequestException as e:
        log.error("eBay API error for '%s': %s", card_name, e)
        return None

    parsed_items = []
    for raw in items:
        parsed = parse_active_item(raw, card_name)
        if parsed:
            parsed_items.append(parsed)

    time.sleep(0.5)

    prices = [p["price"] for p in parsed_items if p.get("currency") == "USD"]
    if not prices:
        return None

    sorted_prices = sorted(prices)
    top5 = sorted_prices[:5]

    snapshot = {
        "card_query":    card_name,
        "snapshot_date": datetime.now(timezone.utc).date().isoformat(),
        "sample_size":   len(top5),
        "avg_price":     round(mean(top5), 2),
        "min_price":     round(min(top5), 2),
        "max_price":     round(max(top5), 2),
    }

    save_active_snapshot(conn, snapshot)
    return snapshot


def ensure_active_price_columns(ws, headers: list[str]) -> dict[str, int]:
    col_map = {}
    last_data_col = 0
    for cell in ws[1]:
        if cell.value is not None:
            name = str(cell.value).strip()
            col_map[name] = cell.column - 1
            if not name.startswith("Last updated"):
                last_data_col = cell.column

    next_col = last_data_col + 1
    price_col_map = {}
    for col_name, _ in ACTIVE_PRICE_COLUMNS:
        if col_name in col_map:
            price_col_map[col_name] = col_map[col_name]
        else:
            cell = ws.cell(row=1, column=next_col, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            price_col_map[col_name] = next_col - 1
            next_col += 1

    return price_col_map


def write_price_data(ws, ws_rows: list[dict], headers: list[str],
                     price_col_map: dict[str, int],
                     card_prices: dict[str, dict | None]) -> None:
    for row in ws_rows:
        row_idx = row["_row"]
        card = row.get("Card")
        if not card or not str(card).strip():
            continue
        card = str(card).strip()
        snapshot = card_prices.get(card)

        for col_name, fmt in ACTIVE_PRICE_COLUMNS:
            col_idx = price_col_map.get(col_name)
            if col_idx is None:
                continue

            cell = ws.cell(row=row_idx, column=col_idx + 1)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")

            if not snapshot:
                cell.value = None
                continue

            if col_name == "Active Avg (Top 5)":
                cell.value = snapshot["avg_price"]
                if fmt:
                    cell.number_format = fmt


def main():
    args = parse_args()
    db_path = args.db

    xlsx_path = args.output
    if not os.path.exists(xlsx_path):
        sys.exit(1)

    wb = load_workbook(xlsx_path)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(1)

    ws = wb[SHEET_NAME]
    ws_rows, headers, col_map = read_active_listings(ws)

    if len(ws_rows) > args.max_listings:
        print(f"Limiting to {args.max_listings} of {len(ws_rows)} rows")
        ws_rows = ws_rows[:args.max_listings]

    if not ws_rows:
        sys.exit(0)

    unique_cards = collect_unique_cards(ws_rows, headers, col_map)
    print(f"Found {len(unique_cards)} cards — searching", end="", flush=True)

    if not unique_cards:
        sys.exit(0)

    conn = init_db(db_path)
    try:
        card_prices: dict[str, dict | None] = {}
        found = 0
        for card in unique_cards:
            try:
                snapshot = fetch_active_price_for_card(conn, card)
            except Exception as e:
                log.error("Unexpected error for '%s': %s", card, e)
                snapshot = None
            card_prices[card] = snapshot
            if snapshot:
                found += 1

        price_col_map = ensure_active_price_columns(ws, headers)
        write_price_data(ws, ws_rows, headers, price_col_map, card_prices)

        wb.save(xlsx_path)
        print(f"\n  {found}/{len(unique_cards)} cards had active listing data")
        print(f"Active price columns saved to {xlsx_path}")

        print("\n--- Active Price Summary ---")
        all_avgs = []
        for card, snap in sorted(card_prices.items()):
            if snap:
                print(f"  {card:<50s}  ${snap['avg_price']:>6.2f}  (n={snap['sample_size']})")
                all_avgs.append(snap["avg_price"])
            else:
                print(f"  {card:<50s}  {'—':>8s}")
        if all_avgs:
            grand = mean(all_avgs)
            print(f"  {'—'*72}")
            print(f"  {'Grand Avg (all cards):':<50s}  ${grand:>6.2f}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
