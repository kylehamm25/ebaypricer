import argparse
import logging
import os
import sqlite3
import sys
import time
from datetime import datetime, timezone

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from ebaypricer.browse_api import (
    compute_snapshot,
    get_today_snapshot,
    init_db,
    parse_item,
    search_sold_listings,
)
from ebaypricer.excel import HEADER_FILL, HEADER_FONT, DATA_FONT
from ebaypricer.paths import DB_PATH as DEFAULT_DB_PATH

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

SHEET_NAME = "Active Listings"
LOOKBACK_DAYS = 30
MAX_LISTINGS = 9999
MAX_SOLD_MATCHES = 10
DEFAULT_OUTPUT = r"H:\My Drive\ebay\ebay_sold_orders.xlsx"

PRICE_COLUMNS = [
    ("Recent Sold Avg",   '#,##0.00'),
    ("Price vs Sold Avg", '#,##0.00'),
    ("Recent Sold Count",  '0'),
    ("Last Checked",      None),
]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Add recent sold-price columns to Active Listings sheet"
    )
    parser.add_argument("--output", type=str, default=DEFAULT_OUTPUT,
                        help="Path to the Excel workbook")

    parser.add_argument("--db", type=str, default=DEFAULT_DB_PATH,
                        help="Path to SQLite database")
    parser.add_argument("--max-listings", type=int, default=MAX_LISTINGS,
                        help="Max number of listings to process")
    return parser.parse_args()


def read_active_listings(ws) -> tuple[list[dict], list[str], dict[str, int]]:
    headers = [cell.value for cell in ws[1]]
    col_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_map[str(h).strip()] = i

    card_col = col_map.get("Card")
    if card_col is None:
        sys.exit(0)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) == 0 or row[0] is None:
            continue
        rows.append({headers[i] if i < len(headers) else None: row[i] for i in range(len(row))})

    return rows, headers, col_map


def collect_unique_cards(rows: list[dict], card_col_idx: int, headers: list[str]) -> list[str]:
    seen: set[str] = set()
    cards: list[str] = []
    for row in rows:
        val = row.get(headers[card_col_idx]) if card_col_idx < len(headers) else row.get("Card")
        if val and str(val).strip():
            card = str(val).strip()
            if card not in seen:
                seen.add(card)
                cards.append(card)
    return cards


def fetch_price_for_card(conn, card_name: str) -> dict | None:
    snapshot = get_today_snapshot(conn, card_name)
    if snapshot:
        return snapshot

    print(".", end="", flush=True)
    try:
        items = search_sold_listings(card_name, LOOKBACK_DAYS)
    except requests.HTTPError as e:
        log.error("eBay API error for '%s': %s", card_name, e)
        return None

    items.sort(
        key=lambda i: i.get("itemEndDate") or i.get("itemCreationDate") or "",
        reverse=True,
    )
    items = items[:MAX_SOLD_MATCHES]

    for raw in items:
        parsed = parse_item(raw, card_name)
        if not parsed:
            continue
        try:
            conn.execute(
                """
                INSERT OR IGNORE INTO sold_listings
                    (item_id, card_query, title, price, currency, condition,
                     listing_type, sold_date, url, pulled_at)
                VALUES
                    (:item_id, :card_query, :title, :price, :currency, :condition,
                     :listing_type, :sold_date, :url, :pulled_at)
                """,
                parsed,
            )
        except Exception:
            pass
    conn.commit()
    time.sleep(1)

    snapshot = compute_snapshot(conn, card_name, LOOKBACK_DAYS)
    return snapshot


def remove_orphan_columns(ws, headers: list[str]) -> list[str]:
    price_names = {name for name, _ in PRICE_COLUMNS}
    to_remove = []
    for i, h in enumerate(headers):
        if h and h.startswith("Recent Sold ") and h not in price_names:
            to_remove.append(i)
    for idx in reversed(to_remove):
        ws.delete_cols(idx + 1)
        del headers[idx]
    return headers


def ensure_price_columns(ws, headers: list[str]) -> dict[str, int]:
    col_map = {}
    last_data_col = 0
    for i, h in enumerate(headers):
        if h is not None:
            col_map[str(h).strip()] = i
            last_data_col = i + 1

    next_col = last_data_col + 1
    price_col_map = {}
    for col_name, _ in PRICE_COLUMNS:
        if col_name in col_map:
            price_col_map[col_name] = col_map[col_name]
        else:
            cell = ws.cell(row=1, column=next_col, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_map[col_name] = next_col - 1
            price_col_map[col_name] = next_col - 1
            next_col += 1

    return price_col_map


def write_price_data(ws, ws_rows: list[dict], headers: list[str],
                     price_col_map: dict[str, int], card_prices: dict[str, dict | None]) -> None:
    for row_idx, row in enumerate(ws_rows, 2):
        card = row.get("Card")
        if not card or not str(card).strip():
            continue
        card = str(card).strip()
        snapshot = card_prices.get(card)

        for col_name, fmt in PRICE_COLUMNS:
            col_idx = price_col_map.get(col_name)
            if col_idx is None:
                continue

            cell = ws.cell(row=row_idx, column=col_idx + 1)
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="center")

            if not snapshot:
                cell.value = None
                continue

            if col_name == "Recent Sold Avg":
                cell.value = snapshot["weighted_avg"]
                if fmt:
                    cell.number_format = fmt
            elif col_name == "Price vs Sold Avg":
                price = row.get("Price")
                if price is not None:
                    try:
                        cell.value = round(float(price) - snapshot["weighted_avg"], 2)
                    except (TypeError, ValueError):
                        cell.value = None
                if fmt:
                    cell.number_format = fmt
            elif col_name == "Recent Sold Count":
                cell.value = snapshot["sample_size"]
                if fmt:
                    cell.number_format = fmt
            elif col_name == "Last Checked":
                cell.value = snapshot["snapshot_date"]


def main():
    args = parse_args()
    db_path = args.db

    today = datetime.now(timezone.utc).date().isoformat()
    if os.path.exists(db_path):
        try:
            conn = sqlite3.connect(db_path)
            count = conn.execute(
                "SELECT COUNT(*) FROM price_snapshots WHERE snapshot_date = ?",
                (today,),
            ).fetchone()[0]
            conn.close()
            if count > 0:
                print(f"Snapshots exist for {today} — skipping")
                return
        except Exception:
            pass

    xlsx_path = args.output

    if not os.path.exists(xlsx_path):
        sys.exit(1)

    wb = load_workbook(xlsx_path)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(1)

    ws = wb[SHEET_NAME]
    ws_rows, headers, col_map = read_active_listings(ws)
    headers = remove_orphan_columns(ws, headers)

    if len(ws_rows) > args.max_listings:
        print(f"Limiting to {args.max_listings} of {len(ws_rows)} rows")
        ws_rows = ws_rows[:args.max_listings]

    if not ws_rows:
        sys.exit(0)

    unique_cards = collect_unique_cards(ws_rows, col_map.get("Card", -1), headers)
    print(f"Found {len(unique_cards)} cards — searching", end="", flush=True)

    if not unique_cards:
        sys.exit(0)

    conn = init_db(db_path)
    try:
        card_prices: dict[str, dict | None] = {}
        found = 0
        for card in unique_cards:
            snapshot = fetch_price_for_card(conn, card)
            card_prices[card] = snapshot
            if snapshot:
                found += 1

        price_col_map = ensure_price_columns(ws, headers)
        write_price_data(ws, ws_rows, headers, price_col_map, card_prices)

        wb.save(xlsx_path)
        print(f"\n  {found}/{len(unique_cards)} cards had sold data")
        print(f"Price columns saved to {xlsx_path}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
