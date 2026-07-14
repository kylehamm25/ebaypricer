import argparse
import os
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ebaypricer.auth import get_access_token
from ebaypricer.trading_api import fetch_active_listings
from ebaypricer.cards import enrich_rows
from ebaypricer.marketing_api import get_campaigns, get_ads
from ebaypricer.excel import (
    HEADER_FILL, HEADER_FONT, DATA_FONT, SHADE_FILL,
    ACTIVE_CURRENCY_COLS, ACTIVE_INT_COLS, write_headers,
)

SHEET_NAME = "Active Listings"

DEFAULT_OUTPUT = r"H:\My Drive\ebay\ebay_sold_orders.xlsx"


def parse_args():
    parser = argparse.ArgumentParser(description="Refresh Active Listings sheet with current eBay listings")
    parser.add_argument("--output", type=str, default=DEFAULT_OUTPUT, help="Output xlsx path")
    return parser.parse_args()


def _clear_sheet(ws: Worksheet) -> None:
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)


def write_data_rows(ws: Worksheet, rows: list[dict], headers: list[str], start_row: int = 2) -> None:
    for row_idx, row in enumerate(rows, start_row):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(h))
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if h in ACTIVE_CURRENCY_COLS and row.get(h) is not None:
                cell.number_format = '#,##0.00'
            elif h in ACTIVE_INT_COLS:
                cell.number_format = '0'
            elif h == "Net Return %" and row.get(h) is not None:
                cell.number_format = '0.00%'
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = SHADE_FILL


def auto_column_widths(ws: Worksheet, headers: list[str], rows: list[dict]) -> None:
    for col_idx, h in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(h)),
            max((len(str(row.get(h, ""))) for row in rows), default=0),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def write_last_updated(ws: Worksheet, now: datetime) -> None:
    cell = ws.cell(row=1, column=50, value=f"Last updated: {now.strftime('%Y-%m-%d %H:%M UTC')}")
    cell.font = Font(italic=True, name="Arial", size=9, color="808080")


def main():
    args = parse_args()
    now = datetime.now(timezone.utc)

    xlsx_path = args.output

    PRICE_COLS_TO_SAVE = ["Recent Sold Avg", "Price vs Sold Avg", "Recent Sold Count", "Last Checked"]

    ws = None
    existing_cards: dict[str, str] = {}
    existing_prices: dict[str, dict[str, object]] = {}
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
            col_to_idx = {str(cell.value): i for i, cell in enumerate(ws[1]) if cell.value}
            card_col = col_to_idx.get("Card", -1)
            itemid_col = col_to_idx.get("Item ID", -1)
            if card_col >= 0 and itemid_col >= 0:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    item_id = str(row[itemid_col]).strip() if len(row) > itemid_col and row[itemid_col] else ""
                    card_val = str(row[card_col]).strip() if len(row) > card_col and row[card_col] else ""
                    if item_id and card_val:
                        existing_cards[item_id] = card_val
            price_cols_map: dict[str, int] = {}
            for name in PRICE_COLS_TO_SAVE:
                idx = col_to_idx.get(name)
                if idx is not None:
                    price_cols_map[name] = idx
            if price_cols_map and itemid_col is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    item_id = str(row[itemid_col]).strip() if len(row) > itemid_col and row[itemid_col] else ""
                    if not item_id:
                        continue
                    prices: dict[str, object] = {}
                    for name, idx in price_cols_map.items():
                        if len(row) > idx:
                            prices[name] = row[idx]
                    if any(v is not None for v in prices.values()):
                        existing_prices[item_id] = prices

    else:
        wb = Workbook()
        default_ws = wb.active
        if default_ws is not None:
            wb.remove(default_ws)

    token = get_access_token()

    rows = fetch_active_listings(token)

    if not rows:
        print("No active listings returned — leaving existing sheet untouched.")
        return

    SHIPPING_PRICE_MAP = {
        "free ebay standard": 0,
        "ebay standard envelope": 0.78,
        "ground advantage": 5,
        "free ground advantage": 0,
    }

    for row in rows:
        profile = (row.get("Shipping Profile") or "").strip().lower()
        row["Shipping Charge"] = SHIPPING_PRICE_MAP.get(profile, 0)

    enrich_rows(rows, title_key="Title")

    if existing_cards:
        for row in rows:
            card = existing_cards.get(row.get("Item ID", ""))
            if card:
                row["Card"] = card

    try:
        campaigns = get_campaigns(token)
        ad_rate_by_listing: dict[str, float] = {}
        if campaigns:
            ads = get_ads(token, campaigns[0]["campaignId"])
            for ad in ads:
                lid = ad.get("listingId", "")
                try:
                    ad_rate_by_listing[lid] = float(ad.get("bidPercentage") or 0)
                except (ValueError, TypeError):
                    pass
        for row in rows:
            rate = ad_rate_by_listing.get(row.get("Item ID", ""))
            if rate is not None:
                row["Ad Rate"] = f"{rate:.0f}%"
        if ad_rate_by_listing:
            print(f"  {len(ad_rate_by_listing)} promoted")
    except SystemExit:
        pass
    except Exception as e:
        print(f"  Ads: {e}")

    COLUMN_ORDER = [
        "Item ID", "Title", "Card", "SKU", "Price",
        "Shipping Charge", "Ad Rate", "Watchers", "Days Listed",
        "Start Date", "Quantity", "Estimated Fees",
        "Estimated Net",
    ]

    for row in rows:
        price = 0.0
        try:
            price = float(row.get("Price") or 0)
        except (TypeError, ValueError):
            pass

        if price <= 2:
            multiplier = 0.65
        elif price <= 5:
            multiplier = 0.70
        else:
            multiplier = 0.73

        estimated_net = round(price * multiplier, 2)
        estimated_fees = round(price - estimated_net, 2)

        row["Estimated Fees"] = estimated_fees
        row["Estimated Net"] = estimated_net

    rows = [{k: row[k] for k in COLUMN_ORDER if k in row} for row in rows]

    rows.sort(key=lambda r: r.get("Days Listed", 0), reverse=True)
    headers = list(rows[0].keys())

    if ws is None:
        ws = wb.create_sheet(SHEET_NAME)
    _clear_sheet(ws)

    write_headers(ws, headers)
    ws.freeze_panes = "A2"
    write_data_rows(ws, rows, headers)

    if existing_prices:
        data_headers = list(rows[0].keys())
        price_start_col = len(data_headers) + 1
        for col_name in PRICE_COLS_TO_SAVE:
            cell = ws.cell(row=1, column=price_start_col, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            price_start_col += 1
        for row_idx, row in enumerate(rows, 2):
            item_id = row.get("Item ID", "")
            saved = existing_prices.get(item_id)
            if not saved:
                continue
            col_offset = len(data_headers) + 1
            for col_name in PRICE_COLS_TO_SAVE:
                val = saved.get(col_name)
                if val is not None:
                    ws.cell(row=row_idx, column=col_offset, value=val)  # type: ignore
                col_offset += 1
        print(f"  Restored {len(existing_prices)} price rows")

    auto_column_widths(ws, headers, rows)
    if existing_prices:
        for col_idx, h in enumerate(PRICE_COLS_TO_SAVE, len(headers) + 1):
            max_len = max(
                len(str(h)),
                max((len(str(v.get(h, ""))) for v in existing_prices.values()), default=0),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 20)

    write_last_updated(ws, now)

    wb.save(xlsx_path)
    print(f"Active Listings refreshed — {len(rows)} listings")


if __name__ == "__main__":
    main()
