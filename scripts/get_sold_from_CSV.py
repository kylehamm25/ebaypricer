"""
Usage:
    python get_sold_from_CSV.py
    python get_sold_from_CSV.py --days 90
    python get_sold_from_CSV.py --start 2025-01-01 --end 2025-06-30
    python get_sold_from_CSV.py --output my_sales.xlsx
    python get_sold_from_CSV.py --debug-fees
    python get_sold_from_CSV.py --skip-fees
"""

import argparse
import csv
import json
import os
import sys
from datetime import datetime, timedelta, timezone

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from sold_api import (
    NS, get_access_token, fetch_sold_orders, _parse_usd, _parse_csv_date,
)

_env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env")
load_dotenv(dotenv_path=_env_path)

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FINANCE_URL = "https://apiz.ebay.com/sell/finances/v1/transaction"

def read_csv_orders(csv_path: str) -> list[dict]:
    if not os.path.isabs(csv_path):
        csv_path = os.path.join(_SCRIPT_DIR, csv_path)
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader)
        headers = next(reader)

        col = {h: i for i, h in enumerate(headers)}
        needed = {"Order Number", "Item Number", "Quantity", "Sold For",
                  "Shipping And Handling", "Total Price", "Sale Date",
                  "Buyer Username", "Item Title", "Custom Label"}
        missing = needed - set(col)
        if missing:
            print(f"ERROR: CSV missing columns: {', '.join(sorted(missing))}")
            sys.exit(1)

        rows = []
        summary = {}
        max_col = max(col.values())
        for csv_row in reader:
            if not csv_row or len(csv_row) <= max_col or not csv_row[col["Order Number"]].strip():
                continue
            order_num = csv_row[col["Order Number"]].strip()
            item_num = csv_row[col["Item Number"]].strip()

            if not item_num:
                summary[order_num] = {
                    "shipping": _parse_usd(csv_row[col["Shipping And Handling"]]),
                    "total": _parse_usd(csv_row[col["Total Price"]]),
                }
                continue

            qty = int(csv_row[col["Quantity"]].strip() or "1")
            sold_for = _parse_usd(csv_row[col["Sold For"]])
            shipping = _parse_usd(csv_row[col["Shipping And Handling"]])
            total_price = _parse_usd(csv_row[col["Total Price"]])

            if order_num in summary:
                if shipping == 0.0 and summary[order_num]["shipping"] != 0.0:
                    shipping = summary[order_num]["shipping"]
                if total_price == 0.0 and summary[order_num]["total"] != 0.0:
                    total_price = summary[order_num]["total"]

            rows.append({
                "Order ID": order_num,
                "Item ID": item_num,
                "Sale Date": _parse_csv_date(csv_row[col["Sale Date"]]),
                "Buyer": csv_row[col["Buyer Username"]].strip(),
                "Item Title": csv_row[col["Item Title"]].strip(),
                "SKU": csv_row[col["Custom Label"]].strip(),
                "Quantity": qty,
                "Item Price": sold_for,
                "Subtotal": round(sold_for * qty, 2),
                "Shipping": shipping,
                "Order Total": total_price,
            })

    if not rows:
        print("ERROR: No order rows found in CSV.")
        sys.exit(1)

    print(f"Read {len(rows)} line items from CSV")
    return rows

def fetch_finance_fees(access_token: str, start_dt: datetime, end_dt: datetime, debug: bool = False):
    """
    Returns:
        fees_by_order: {real_order_id: {feeType: amount}}
        item_id_index: {item_id: [(transaction_date_iso, real_order_id), ...]}
        fee_types:      sorted list of all distinct feeType strings seen
    """
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
        "X-EBAY-C-MARKETPLACE-ID": "EBAY_US",
    }

    # Widen the window — Finances API "transactionDate" can lag well behind
    # the order creation date Trading API uses (e.g. Promoted Listings fees
    # often post days after the sale), so a tight window risks missing
    # matches. Harmless to over-fetch since merging happens by key/item+date,
    # not by date range.
    pad_start = start_dt - timedelta(days=15)
    pad_end = end_dt + timedelta(days=15)
    date_filter = (
        f"transactionDate:[{pad_start.strftime('%Y-%m-%dT%H:%M:%S.000Z')}.."
        f"{pad_end.strftime('%Y-%m-%dT%H:%M:%S.000Z')}]"
    )

    fees_by_order: dict = {}
    item_id_index: dict = {}
    pending_item_fees: dict = {}  # item_id -> [(feeType, value, date)] for fees with no ORDER_ID ref
    fee_types: set = set()

    url = FINANCE_URL
    params = {"filter": date_filter, "limit": 200}
    page = 1

    while url:
        resp = requests.get(url, headers=headers, params=params if page == 1 else None, timeout=30)
        if resp.status_code != 200:
            print(f"Finances API error ({resp.status_code}): {resp.text[:500]}")
            sys.exit(1)

        data = resp.json()

        if debug and page == 1:
            print("\n--- DEBUG: full Finances API transaction samples ---")
            sale_sample = next((t for t in data.get("transactions", []) if t.get("transactionType") == "SALE"), None)
            other_sample = next((t for t in data.get("transactions", []) if t.get("transactionType") != "SALE"), None)
            print("\n[SALE-type transaction, full]:")
            print(json.dumps(sale_sample, indent=2) if sale_sample else "  (none found on page 1)")
            print("\n[Non-SALE-type transaction, full]:")
            print(json.dumps(other_sample, indent=2) if other_sample else "  (none found on page 1)")
            print("--- end debug ---\n")

        transactions = data.get("transactions", [])
        print(f"  Finances page {page} — {len(transactions)} transactions")

        for txn in transactions:
            tdate = txn.get("transactionDate", "")

            if txn.get("transactionType") == "SALE":
                order_id = txn.get("orderId", "")
                for li in txn.get("orderLineItems", []):
                    iid = li.get("legacyItemId") or li.get("itemId")
                    liid = li.get("lineItemId", "")
                    if not iid and liid:
                        iid = liid
                    if iid and order_id:
                        item_id_index.setdefault(iid, []).append((tdate, order_id))
                    if liid and liid != iid and order_id:
                        item_id_index.setdefault(liid, []).append((tdate, order_id))

                    for fee in li.get("marketplaceFees", []):
                        fee_type = fee.get("feeType", "UNKNOWN_FEE")
                        try:
                            value = abs(float(fee.get("amount", {}).get("value", 0.0)))
                        except (TypeError, ValueError):
                            value = 0.0
                        fee_types.add(fee_type)
                        bucket = fees_by_order.setdefault(order_id, {})
                        bucket[fee_type] = bucket.get(fee_type, 0.0) + value

            elif txn.get("feeType"):
                fee_type = txn.get("feeType", "UNKNOWN_FEE")
                try:
                    value = abs(float(txn.get("amount", {}).get("value", 0.0)))
                except (TypeError, ValueError):
                    value = 0.0
                refs = txn.get("references", [])
                order_ref = next((r.get("referenceId") for r in refs if r.get("referenceType") == "ORDER_ID"), None)
                item_ref = next((r.get("referenceId") for r in refs if r.get("referenceType") == "ITEM_ID"), None)

                fee_types.add(fee_type)
                if order_ref:
                    bucket = fees_by_order.setdefault(order_ref, {})
                    bucket[fee_type] = bucket.get(fee_type, 0.0) + value
                elif item_ref:
                    pending_item_fees.setdefault(item_ref, []).append((fee_type, value, tdate))

        url = data.get("next")
        params = None  # "next" already contains the query string
        page += 1
        if page > 100:  # sanity guard against unexpected infinite pagination
            print("WARNING: stopped fee pagination after 100 pages — check filter/limit.")
            break

    # Resolve any item-only-referenced fees (no ORDER_ID ref) against the
    # item_id_index built from SALE transactions, picking the closest date.
    for item_id, entries in pending_item_fees.items():
        candidates = item_id_index.get(item_id, [])
        if not candidates:
            continue
        for fee_type, value, fee_date in entries:
            real_order_id = _closest_by_date(candidates, fee_date)
            if real_order_id:
                bucket = fees_by_order.setdefault(real_order_id, {})
                bucket[fee_type] = bucket.get(fee_type, 0.0) + value

    return fees_by_order, item_id_index


def _closest_by_date(candidates: list, target_date_str: str):
    """candidates: list of (date_iso_str, real_order_id). Returns the
    real_order_id whose date is closest to target_date_str, or the only
    candidate if there's just one."""
    if len(candidates) == 1:
        return candidates[0][1]
    try:
        target = datetime.fromisoformat(target_date_str.replace("Z", "+00:00"))
    except (ValueError, AttributeError):
        return candidates[0][1]

    best_id, best_diff = None, None
    for date_str, order_id in candidates:
        try:
            d = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
        except (ValueError, AttributeError):
            continue
        diff = abs((d - target).total_seconds())
        if best_diff is None or diff < best_diff:
            best_diff, best_id = diff, order_id
    return best_id or candidates[0][1]


def merge_fees_into_rows(rows: list[dict], fees_by_order: dict, item_id_index: dict) -> None:
    if not fees_by_order:
        for row in rows:
            row["Total eBay Fees"] = None
            row["Order Earnings"] = None
        return

    groups: dict = {}
    for row in rows:
        groups.setdefault(row["Order ID"], []).append(row)

    for trading_order_id, group in groups.items():
        real_order_id = trading_order_id if trading_order_id in fees_by_order else None

        if real_order_id is None:
            for row in group:
                iid = row.get("Item ID")
                sale_date = row.get("Sale Date")

                if iid and sale_date and iid in item_id_index:
                    candidates = item_id_index[iid]
                    real_order_id = _closest_by_date(candidates, sale_date)
                    if real_order_id:
                        break

                oid = row.get("Order ID", "")
                if "-" in oid and sale_date:
                    liid = oid.split("-", 1)[1]
                    if liid in item_id_index:
                        candidates = item_id_index[liid]
                        real_order_id = _closest_by_date(candidates, sale_date)
                        if real_order_id:
                            break

        fees = fees_by_order.get(real_order_id) if real_order_id else None
        total_fees = round(sum(fees.values()), 2) if fees else None

        for row in group:
            row["Total eBay Fees"] = total_fees
            order_total = row.get("Order Total") or 0.0
            shipping = row.get("Shipping") or 0.0
            row["Order Earnings"] = round(order_total - total_fees - shipping, 2) if total_fees is not None else None


# ── Combine Orders ───────────────────────────────────────────────────────────

def combine_orders(rows: list[dict]) -> list[dict]:
    """Group rows by Order ID, combining multi-item orders into one row."""
    groups: dict = {}
    order = []
    for row in rows:
        oid = row["Order ID"]
        if oid not in groups:
            groups[oid] = []
            order.append(oid)
        groups[oid].append(row)

    combined = []
    for oid in order:
        group = groups[oid]
        if len(group) == 1:
            combined.append(group[0])
            continue

        first = group[0]
        skus = [r["SKU"] for r in group if r["SKU"]]
        links = [r["Link"] for r in group if r.get("Link")]
        combined.append({
            "Order ID": oid,
            "Item ID": "; ".join(r["Item ID"] for r in group),
            "Sale Date": first["Sale Date"],
            "Buyer": first["Buyer"],
            "Item Title": "; ".join(r["Item Title"] for r in group),
            "SKU": "; ".join(skus),
            "Quantity": sum(r["Quantity"] for r in group),
            "Item Price": first["Item Price"],
            "Subtotal": sum(r["Subtotal"] for r in group),
            "Shipping": first["Shipping"],
            "Order Total": first["Order Total"],
            "Link": "; ".join(links),
        })
        last = combined[-1]
        for key in ("Total eBay Fees", "Order Earnings"):
            last[key] = first.get(key)

    print(f"Combined into {len(combined)} orders ({len(rows) - len(combined)} multi-item collapsed)")
    return combined


# ── Excel ─────────────────────────────────────────────────────────────────────

def write_excel(rows: list[dict], filename: str) -> str:
    wb = Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "Sold Orders"

    headers = list(rows[0].keys()) if rows else []

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    for col_idx, h in enumerate(headers, 1):
        cell: Cell = ws.cell(row=1, column=col_idx, value=h)  # type: ignore[assignment]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_font    = Font(name="Arial", size=10)
    currency_cols = {"Item Price", "Subtotal", "Shipping", "Order Total", "Total eBay Fees", "Order Earnings"}
    int_cols      = {"Quantity"}

    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            cell: Cell = ws.cell(row=row_idx, column=col_idx, value=row[h])  # type: ignore[assignment]
            cell.font = data_font
            cell.alignment = Alignment(vertical="center")
            if h in currency_cols and row[h] is not None:
                cell.number_format = '#,##0.00'
            elif h in int_cols:
                cell.number_format = '0'

        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                shade: Cell = ws.cell(row=row_idx, column=col_idx)  # type: ignore[assignment]
                shade.fill = PatternFill("solid", fgColor="EBF3FB")

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    n = len(rows)
    ws_sum["A1"] = "Summary"
    ws_sum["A1"].font = Font(bold=True, name="Arial", size=12)

    col_letter = {h: get_column_letter(i) for i, h in enumerate(headers, 1)}

    ws_sum["A3"] = "Total Transactions"
    ws_sum["B3"] = f"=COUNTA('Sold Orders'!{col_letter['Order ID']}2:{col_letter['Order ID']}{n + 1})"
    ws_sum["A4"] = "Total Revenue"
    ws_sum["B4"] = f"=SUM('Sold Orders'!{col_letter['Order Total']}2:{col_letter['Order Total']}{n + 1})"
    ws_sum["B4"].number_format = '#,##0.00'
    ws_sum["A5"] = "Total Shipping Collected"
    ws_sum["B5"] = f"=SUM('Sold Orders'!{col_letter['Shipping']}2:{col_letter['Shipping']}{n + 1})"
    ws_sum["B5"].number_format = '#,##0.00'

    next_row = 6
    if "Total eBay Fees" in col_letter:
        ws_sum[f"A{next_row}"] = "Total eBay Fees"
        ws_sum[f"B{next_row}"] = (
            f"=SUM('Sold Orders'!{col_letter['Total eBay Fees']}2:"
            f"{col_letter['Total eBay Fees']}{n + 1})"
        )
        ws_sum[f"B{next_row}"].number_format = '#,##0.00'
        next_row += 1
    if "Order Earnings" in col_letter:
        ws_sum[f"A{next_row}"] = "Total Order Earnings"
        ws_sum[f"B{next_row}"] = (
            f"=SUM('Sold Orders'!{col_letter['Order Earnings']}2:"
            f"{col_letter['Order Earnings']}{n + 1})"
        )
        ws_sum[f"B{next_row}"].number_format = '#,##0.00'
        next_row += 1

    ws_sum[f"A{next_row}"] = "Avg Order Value"
    ws_sum[f"B{next_row}"] = "=IF(B3=0,0,B4/B3)"
    ws_sum[f"B{next_row}"].number_format = '#,##0.00'
    last_row = next_row

    for r in range(3, last_row + 1):
        ws_sum.cell(r, 1).font = Font(name="Arial", size=10, bold=True)
        ws_sum.cell(r, 2).font = Font(name="Arial", size=10)

    note_row = last_row + 2
    notes = []
    if "Order Earnings" in col_letter:
        notes.append(
            "Note: 'Order Earnings' = Order Total minus eBay fees. It does not "
            "subtract refunds (not tracked here) or your own item cost — eBay's "
            "Seller Hub 'Net order earnings' figure additionally subtracts a "
            "per-item cost you enter manually in Seller Hub, which isn't "
            "exposed via any API."
        )
    if any(row.get("Total eBay Fees") is None for row in rows):
        notes.append(
            "Note: some rows show blank fees because no matching transaction "
            "was found in the Finances API response for that order/item within "
            "the date range searched — verify with --debug-fees if unexpected."
        )
    for note in notes:
        ws_sum[f"A{note_row}"] = note
        ws_sum[f"A{note_row}"].font = Font(name="Arial", size=9, italic=True, color="808080")
        note_row += 1

    assert ws is not None
    for col_idx, h in enumerate(headers, 1):
        col_letter_w = get_column_letter(col_idx)
        max_len = max(
            len(str(h)),
            max((len(str(row.get(h, ""))) for row in rows), default=0),
        )
        ws.column_dimensions[col_letter_w].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    wb.save(filename)
    return filename


# ── CLI ───────────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(description="Export eBay sold orders to Excel via Trading API")
    parser.add_argument("--csv",    type=str,                                help="Read orders from CSV instead of Trading API")
    parser.add_argument("--days",   type=int, default=30,                    help="Past days to fetch (default: 30)")
    parser.add_argument("--start",  type=str,                                help="Start date YYYY-MM-DD (overrides --days)")
    parser.add_argument("--end",    type=str,                                help="End date YYYY-MM-DD (default: today)")
    parser.add_argument("--output", type=str, default=r"H:\My Drive\ebay\ebay_sold_orders.xlsx", help="Output filename")
    parser.add_argument("--skip-fees", action="store_true", help="Skip the Finances API fee/earnings lookup entirely")
    parser.add_argument("--debug-fees", action="store_true", help="Print raw Finances API JSON samples for verifying field names")
    return parser.parse_args()


def main():
    args = parse_args()

    now    = datetime.now(timezone.utc)
    end_dt = datetime.strptime(args.end, "%Y-%m-%d").replace(tzinfo=timezone.utc) if args.end else now
    start_dt = (
        datetime.strptime(args.start, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        if args.start
        else end_dt - timedelta(days=args.days)
    )

    token = get_access_token()

    if args.csv:
        rows = read_csv_orders(args.csv)
        if not args.start and not args.end:
            dates = [datetime.strptime(r["Sale Date"], "%Y-%m-%d").replace(tzinfo=timezone.utc) for r in rows]
            start_dt = min(dates)
            end_dt = max(dates) + timedelta(days=1)
        print(f"Orders from {start_dt.date()} to {end_dt.date()}")
    else:
        print(f"Fetching sold orders from {start_dt.date()} to {end_dt.date()}...")
        rows = fetch_sold_orders(token, start_dt, end_dt)

    print(f"\nFound {len(rows)} line items.")

    if not rows:
        print("No sold orders found for this date range.")
        sys.exit(0)

    if args.skip_fees:
        for row in rows:
            row["Total eBay Fees"] = None
            row["Order Earnings"] = None
    else:
        print("\nFetching fee/earnings data from Finances API...")
        fees_by_order, item_id_index = fetch_finance_fees(
            token, start_dt, end_dt, debug=args.debug_fees
        )
        print(f"Found fee data for {len(fees_by_order)} orders")
        merge_fees_into_rows(rows, fees_by_order, item_id_index)

    rows = combine_orders(rows)
    output_file = write_excel(rows, args.output)
    print(f"Saved -> {output_file}")


if __name__ == "__main__":
    main()