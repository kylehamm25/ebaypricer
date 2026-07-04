from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT = Font(name="Arial", size=10)
SHADE_FILL = PatternFill("solid", fgColor="EBF3FB")

CURRENCY_COLS = {"Item Price", "Shipping", "Order Total", "Total eBay Fees", "Order Earnings"}
INT_COLS = {"Quantity"}
ACTIVE_CURRENCY_COLS = {"Price"}
ACTIVE_INT_COLS = {"Quantity", "Days Listed", "Watchers"}


def write_headers(ws, headers):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_data_rows(ws, rows, headers, start_row=2, currency_cols=None, int_cols=None):
    if currency_cols is None:
        currency_cols = CURRENCY_COLS
    if int_cols is None:
        int_cols = INT_COLS
    for row_idx, row in enumerate(rows, start_row):
        for col_idx, h in enumerate(headers, 1):
            val = row.get(h)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="center")
            if h in currency_cols and val is not None:
                cell.number_format = '#,##0.00'
            elif h in int_cols:
                cell.number_format = '0'
